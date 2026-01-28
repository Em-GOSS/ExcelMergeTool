import os
import subprocess
import sys
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import Series, SeriesLabel


HEADERS = [
    "P",
    "FL",
    "FR",
    "RL",
    "RR",
    "Average_Front",
    "Average_Rear",
]
BLOCK_WIDTH = 8  # 7 columns + 1 spacer


def read_data_unit(path):
    workbook = load_workbook(path)
    sheet = workbook.active
    filename = os.path.splitext(os.path.basename(path))[0]

    headers = []
    for col in range(1, 8):
        headers.append(sheet.cell(row=1, column=col).value)

    data_rows = []
    row = 2
    while True:
        row_values = [sheet.cell(row=row, column=col).value for col in range(1, 8)]
        if all(value in (None, "") for value in row_values):
            break
        data_rows.append(row_values)
        row += 1

    return filename, headers, data_rows


def find_next_block_column(sheet):
    col = 1
    while sheet.cell(row=1, column=col).value not in (None, ""):
        col += BLOCK_WIDTH
    return col


def append_data_unit(sheet, filename, headers, data_rows):
    start_col = find_next_block_column(sheet)
    sheet.cell(row=1, column=start_col, value=filename)

    for offset, header in enumerate(headers):
        sheet.cell(row=2, column=start_col + offset, value=header)

    start_row = 3
    for row_index, row_values in enumerate(data_rows):
        for col_offset, value in enumerate(row_values):
            sheet.cell(row=start_row + row_index, column=start_col + col_offset, value=value)


def get_block_columns(sheet):
    block_cols = []
    max_col = sheet.max_column
    col = 1
    while col <= max_col:
        if sheet.cell(row=1, column=col).value not in (None, ""):
            block_cols.append(col)
        col += BLOCK_WIDTH
    return block_cols


def find_last_data_row(sheet, start_col):
    row = 3
    while sheet.cell(row=row, column=start_col).value not in (None, ""):
        row += 1
    return row - 1


def build_chart(title, y_col_offset, sheet, block_cols):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = title
    chart.x_axis.title = "P"
    chart.width = 30
    chart.height = 15
    chart.legend.position = "b"
    chart.legend.overlay = False

    for index, start_col in enumerate(block_cols):
        filename = sheet.cell(row=1, column=start_col).value
        last_row = find_last_data_row(sheet, start_col)
        if last_row < 3:
            continue
        x_values = Reference(
            sheet,
            min_col=start_col,
            min_row=3,
            max_row=last_row,
        )
        y_values = Reference(
            sheet,
            min_col=start_col + y_col_offset,
            min_row=3,
            max_row=last_row,
        )
        series = Series(y_values, x_values)
        series.title = SeriesLabel(v=filename)
        series.idx = index
        series.order = index
        chart.series.append(series)
    return chart


def rebuild_charts(workbook, data_sheet):
    if "Charts" in workbook.sheetnames:
        del workbook["Charts"]
    chart_sheet = workbook.create_sheet("Charts")

    block_cols = get_block_columns(data_sheet)
    if not block_cols:
        return

    titles = ["Front Left", "Front Right", "Rear Left", "Rear Right", "Average Front", "Average Rear"]
    offsets = [1, 2, 3, 4, 5, 6]
    positions = ["A1", "T1", "A32", "T32", "A63", "T63"]

    for title, offset, position in zip(titles, offsets, positions):
        chart = build_chart(title, offset, data_sheet, block_cols)
        chart_sheet.add_chart(chart, position)

    workbook.active = workbook.sheetnames.index("Charts")


def open_excel(path):
    if sys.platform.startswith("win"):
        os.startfile(path)
        return
    if sys.platform == "darwin":
        subprocess.run(["open", path], check=False)
        return
    subprocess.run(["xdg-open", path], check=False)


def process_files(all_data_path, data_unit_path):
    data_unit_paths = []
    if os.path.isdir(data_unit_path):
        for entry in sorted(os.listdir(data_unit_path)):
            if entry.lower().endswith(".xlsx") and not entry.startswith("~$"):
                data_unit_paths.append(os.path.join(data_unit_path, entry))
        if not data_unit_paths:
            raise ValueError("所选文件夹中没有可用的xlsx文件")
    else:
        data_unit_paths = [data_unit_path]

    workbook = load_workbook(all_data_path)
    data_sheet = workbook.active

    for unit_path in data_unit_paths:
        filename, headers, data_rows = read_data_unit(unit_path)
        if headers != HEADERS:
            raise ValueError("数据单元表头必须是: P, FL, FR, RL, RR, Average_Front, Average_Rear")
        append_data_unit(data_sheet, filename, headers, data_rows)
    rebuild_charts(workbook, data_sheet)

    workbook.save(all_data_path)
    open_excel(all_data_path)


def select_file(entry):
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if path:
        entry.delete(0, tk.END)
        entry.insert(0, path)


def select_folder(entry):
    path = filedialog.askdirectory()
    if path:
        entry.delete(0, tk.END)
        entry.insert(0, path)


def run_gui():
    root = tk.Tk()
    root.title("Excel Merge Tool")

    tk.Label(root, text="AllDataList Excel:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    all_data_entry = tk.Entry(root, width=60)
    all_data_entry.grid(row=0, column=1, padx=5, pady=5)
    tk.Button(root, text="选择", command=lambda: select_file(all_data_entry)).grid(row=0, column=2, padx=5, pady=5)

    tk.Label(root, text="数据单元 Excel 或文件夹:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    data_unit_entry = tk.Entry(root, width=60)
    data_unit_entry.grid(row=1, column=1, padx=5, pady=5)
    tk.Button(root, text="选择文件", command=lambda: select_file(data_unit_entry)).grid(row=1, column=2, padx=5, pady=5)
    tk.Button(root, text="选择文件夹", command=lambda: select_folder(data_unit_entry)).grid(row=1, column=3, padx=5, pady=5)

    def on_process():
        all_data_path = all_data_entry.get().strip()
        data_unit_path = data_unit_entry.get().strip()
        if not all_data_path or not data_unit_path:
            messagebox.showerror("错误", "请选择两个Excel文件")
            return
        try:
            process_files(all_data_path, data_unit_path)
        except Exception as exc:
            messagebox.showerror("处理失败", str(exc))
            return
        messagebox.showinfo("完成", "数据已写入AllDataList并更新图表")

    tk.Button(root, text="开始处理", command=on_process).grid(row=2, column=1, pady=10)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
